"""
Stage 5: Excel output.

For each batch we:
  1. Copy the locked KBC template to a fresh output file (so the master
     template is never touched).
  2. Delete the template's "Example" row (row 4) and input-options hint
     row (row 5) entirely, so the deliverable has no sample data and no
     blank gap above the real data. Because openpyxl does not shift
     data-validation ranges when rows are deleted, we rewrite the dropdown
     ranges ourselves so the locked dropdowns stay aligned.
  3. Walk the records and write them starting at row 4 (right under the
     column headers), then blank any leftover pre-numbered "#" values in
     the rows below so the file ends cleanly after the data.
  4. Auto-number column B (#).
  5. Skip any field that is None so empty cells stay empty (no broken
     formulas, no placeholder text).
  6. Highlight cells the LLM flagged as low-confidence in yellow, so the
     reviewer knows where to focus.

We deliberately do NOT touch rows 1-4: they hold the merged section
headers, the column titles, and the data-validation example. Leaving
them alone preserves all template formatting and dropdowns automatically.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional
import logging
import re
import shutil
import zipfile
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .config import BUILDING_TEMPLATE, LAND_TEMPLATE, Config
from .normalizers import clean_number, normalize_state, normalize_date
from .schemas import BuildingRecord, LandRecord


log = logging.getLogger("flyer_reader")


# Data rows begin at row 5. The source template has the input-options
# hint row at row 5; we clear that row's contents before writing so it
# does not appear in the deliverable. We do NOT call delete_rows()
# because openpyxl does not shift data-validation ranges when rows are
# deleted, which would break the locked dropdowns on every data row.
FIRST_DATA_ROW = 5
# After we delete the example + hint rows from a fresh file, data begins
# here (was row 6 in the template). Append-mode scanning starts here too so
# it works on both the new no-gap layout and any older files.
SCAN_FIRST_ROW = 4
# Row 4 is the template's "Example" row — a fully filled-in sample record
# (column B reads "Example"). It documents the expected format but should
# NOT appear in the deliverable, so we clear it before writing.
EXAMPLE_ROW = 4
# The hint row (was row 5 in source) lives at this row in the output.
# We clear cells in this row in columns A through HINT_ROW_LAST_COL
# before writing the first record. If 0 records are extracted, that's
# fine: row 5 simply ends up blank rather than carrying the hint text.
HINT_ROW = 5
HINT_ROW_LAST_COL_BUILDING = 40  # column AN
HINT_ROW_LAST_COL_LAND = 18      # column R

# Highlight color for low-confidence cells (soft yellow).
LOW_CONFIDENCE_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")


# ---------------------------------------------------------------------------
# Column mappings.
#
# Each entry: pydantic field name -> Excel column letter.
# Order matches the template column order so the file reads top-to-bottom.
# ---------------------------------------------------------------------------

BUILDING_COLUMNS: list[tuple[str, str]] = [
    ("property_name", "C"),
    ("address", "D"),
    ("city", "E"),
    ("state", "F"),
    ("zip_code", "G"),
    ("latitude", "H"),
    ("longitude", "I"),
    ("property_owner", "J"),
    ("total_building_area_sf", "K"),
    ("available_space_sf", "L"),
    ("office_space_sf", "M"),
    ("land_area_acres", "N"),
    ("year_built", "O"),
    ("building_status", "P"),
    ("building_type", "Q"),
    ("occupancy_status", "R"),
    ("tenancy_type", "S"),
    ("current_tenant", "T"),
    ("prior_use", "U"),
    ("prior_tenant", "V"),
    ("date_available", "W"),
    ("sprinkler_type", "X"),
    ("current_zoning", "Y"),
    ("clear_height_min_ft", "Z"),
    ("clear_height_max_ft", "AA"),
    ("load_type", "AB"),
    ("dock_doors", "AC"),
    ("grade_level_doors", "AD"),
    ("column_spacing_width_ft", "AE"),
    ("column_spacing_depth_ft", "AF"),
    ("auto_parking_spaces", "AG"),
    ("trailer_parking_spaces", "AH"),
    ("existing_power", "AI"),
    ("sale_lease", "AJ"),
    ("annual_asking_rate_psf", "AK"),
    ("sale_price", "AL"),
    ("estimated_annual_opex_psf", "AM"),
    ("comments", "AN"),
]

LAND_COLUMNS: list[tuple[str, str]] = [
    ("property_name", "C"),
    ("address", "D"),
    ("city", "E"),
    ("state", "F"),
    ("zip_code", "G"),
    ("latitude", "H"),
    ("longitude", "I"),
    ("land_owner", "J"),
    ("land_area_acres", "K"),
    ("current_zoning", "L"),
    ("land_construction_status", "M"),
    ("existing_power", "N"),
    ("sale_lease", "O"),
    # P (Estimated Sale Price per Acre) is a formula — we write =Q/K so the
    # row's per-acre price stays in sync with sale_price and acreage.
    ("sale_price", "Q"),
    ("comments", "R"),
]


def template_columns(survey_kind: str) -> list[tuple[str, str, str]]:
    """
    Return the template's data columns as (column_letter, header_label,
    field_name) for the given survey kind ("building" or "land"). The
    header_label is the human-readable column title from row 3 of the
    template — handy for user-facing menus.
    """
    if survey_kind == "land":
        template, columns = LAND_TEMPLATE, LAND_COLUMNS
    else:
        template, columns = BUILDING_TEMPLATE, BUILDING_COLUMNS
    from openpyxl.utils import column_index_from_string as _ci
    try:
        ws = load_workbook(template).active
    except Exception:
        # Fall back to the field name if the template can't be read.
        return [(col, field.replace("_", " ").title(), field)
                for field, col in columns]
    out: list[tuple[str, str, str]] = []
    for field, col in columns:
        try:
            hdr = ws.cell(row=3, column=_ci(col)).value
        except Exception:
            hdr = None
        label = str(hdr).strip() if hdr else field.replace("_", " ").title()
        out.append((col, label, field))
    return out


def all_template_columns() -> list[tuple[str, str]]:
    """
    Combined, de-duplicated list of (header_label, field_name) across both
    survey templates, for a survey-agnostic picker (Field Hints presets are
    not tied to a single survey kind). Order: building columns first, then
    any land-only columns. Labels that map to the same field are merged.
    """
    seen: dict[str, str] = {}  # field_name -> label
    order: list[str] = []
    for kind in ("building", "land"):
        for _col, label, field in template_columns(kind):
            if field not in seen:
                seen[field] = label
                order.append(field)
    return [(seen[f], f) for f in order]


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

@dataclass
class WriteSummary:
    output_path: Path
    rows_written: int
    flagged_cells: list[str]   # e.g. ["P6", "AK7"]
    # True when we appended to an existing file rather than creating one.
    appended: bool = False
    # The row at which writing began. For appends this is N+1 (after the
    # last existing record); for fresh files it's FIRST_DATA_ROW.
    start_row: int = FIRST_DATA_ROW


def _strip_external_links(xlsx_path: Path) -> int:
    """
    Sanitise an .xlsx file in place so openpyxl + Excel can't choke on it.

    Why this exists
    ---------------
    Real-world KBC templates carry two kinds of cruft that accumulate when
    a workbook is copied from file to file over years:

    1. External-workbook links — formulas that reference *other* Excel
       files (e.g. ='[Budget.xlsx]Sheet1'!A1), stored as
       `xl/externalLinks/externalLinkN.xml` parts. openpyxl cannot
       round-trip these: on save it drops the externalLink parts but
       keeps the formulas pointing at them, so Excel reports
       "Repaired Records: External formula reference...".

    2. Broken defined names (named ranges) — the templates contain over a
       thousand `<definedName>` entries, virtually all of them dead
       (`#REF!`, `#N/A`, or pointing at the external links above). Once
       the external links are removed, Excel finds these names invalid
       and reports "Removed Records: Named range from /xl/workbook.xml" —
       and during that repair it can blank the whole sheet.

    Neither feature is used by this app (dropdowns use inline lists, the
    only real formula is the per-acre `=Q/K`). So the safe, clean fix is
    to strip BOTH entirely before openpyxl ever opens the file.

    Steps, operating directly on the xlsx zip:
      1. Drop every `xl/externalLinks/*` part.
      2. Drop the matching relationship entries in workbook.xml.rels.
      3. Remove `<externalReferences>` from workbook.xml.
      4. Remove ALL `<definedName>` entries (the `<definedNames>` block)
         from workbook.xml — they are all dead references.
      5. In every worksheet, replace any formula cell whose formula
         references an external link with that cell's cached value.
      6. Remove the matching [Content_Types] overrides.

    Returns the number of external-link parts removed (0 if the file was
    already clean).
    """
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        ext_parts = [n for n in names if n.startswith("xl/externalLinks/")]
        contents = {n: zin.read(n) for n in names}

    wb_name = "xl/workbook.xml"
    wb_xml = contents.get(wb_name, b"").decode("utf-8", "ignore")
    has_defined_names = "<definedName" in wb_xml

    # Nothing to do only if BOTH problems are absent.
    if not ext_parts and not has_defined_names:
        return 0

    # --- 1 & 2. Decide which parts and rel-ids to drop ----------------------
    drop_parts = set(ext_parts)
    rels_name = "xl/_rels/workbook.xml.rels"
    if rels_name in contents:
        rels_xml = contents[rels_name].decode("utf-8", "ignore")
        # Strip Relationship tags whose Target is an externalLink.
        rels_xml = re.sub(
            r"<Relationship\b[^>]*externalLink[^>]*/>", "", rels_xml
        )
        contents[rels_name] = rels_xml.encode("utf-8")

    # --- 3. Remove <externalReferences> from workbook.xml -------------------
    if wb_xml:
        wb_xml = re.sub(
            r"<externalReferences>.*?</externalReferences>", "", wb_xml, flags=re.DOTALL
        )
        wb_xml = re.sub(r"<externalReferences\s*/>", "", wb_xml)

        # --- 4. Remove ALL defined names (named ranges) --------------------
        # The templates carry 1000+ dead named ranges (#REF!, #N/A, and
        # external-link refs). None are used by the app. Excel reports
        # "Removed Records: Named range" and may blank the sheet during
        # repair if any are invalid, so we drop the whole block.
        wb_xml = re.sub(
            r"<definedNames>.*?</definedNames>", "", wb_xml, flags=re.DOTALL
        )
        wb_xml = re.sub(r"<definedNames\s*/>", "", wb_xml)

        contents[wb_name] = wb_xml.encode("utf-8")

    # --- 5. Neutralise formulas that reference an external link ------------
    # External refs look like  [1]Sheet1!A1  inside an <f>...</f> element.
    # We replace the whole <c ...><f>..</f><v>cached</v></c> formula with
    # just the cached <v> value, dropping the <f> entirely.
    def _neutralise_formulas(xml: str) -> str:
        def repl(cm: re.Match) -> str:
            cell = cm.group(0)
            # Only touch cells whose formula references an external link.
            if not re.search(r"<f[^>]*>[^<]*\[\d+\]", cell):
                return cell
            # Keep the cached <v>…</v> if present; else leave the cell empty.
            vmatch = re.search(r"<v>.*?</v>", cell, flags=re.DOTALL)
            open_tag = re.match(r"<c\b[^>]*>", cell).group(0)
            if vmatch:
                return f"{open_tag}{vmatch.group(0)}</c>"
            return f"{open_tag}</c>"

        return re.sub(r"<c\b[^>]*>.*?</c>", repl, xml, flags=re.DOTALL)

    for name in list(contents.keys()):
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
            sheet_xml = contents[name].decode("utf-8", "ignore")
            if "[" in sheet_xml:  # cheap pre-check before the heavy regex
                contents[name] = _neutralise_formulas(sheet_xml).encode("utf-8")

    # --- 6. Remove [Content_Types] entries for the dropped parts -----------
    ct_name = "[Content_Types].xml"
    if ct_name in contents:
        ct_xml = contents[ct_name].decode("utf-8", "ignore")
        ct_xml = re.sub(
            r'<Override\b[^>]*externalLink[^>]*/>', "", ct_xml
        )
        contents[ct_name] = ct_xml.encode("utf-8")

    # --- Rewrite the zip without the dropped parts -------------------------
    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in contents.items():
            if name in drop_parts:
                continue
            # Also skip any rels file living under xl/externalLinks/_rels/.
            if name.startswith("xl/externalLinks/"):
                continue
            zout.writestr(name, data)

    return len(ext_parts)


def _prepare_output(survey_kind: str, output_dir: Path,
                    output_name: Optional[str] = None) -> Path:
    """
    Copy the locked template to a new file inside output_dir.

    If output_name is given, use it verbatim as the filename (caller is
    expected to have already validated characters and added the .xlsx
    extension). Otherwise fall back to the legacy timestamped name:
    KBC_{Building|Land}_Survey_{YYYYMMDD_HHMMSS}.xlsx.
    """
    template = BUILDING_TEMPLATE if survey_kind == "building" else LAND_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(
            f"Template not found: {template}. Check that the templates/ folder "
            f"ships next to the application."
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    if output_name:
        filename = output_name
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"KBC_{survey_kind.capitalize()}_Survey_{stamp}.xlsx"
    out_path = output_dir / filename
    shutil.copy(template, out_path)
    # Defensively strip any external-workbook links so openpyxl can't
    # leave dangling references that make Excel report file corruption.
    removed = _strip_external_links(out_path)
    if removed:
        log.info("Removed %d external-link part(s) from the template copy.", removed)
    return out_path


def detect_target_survey_kind(target_path: Path) -> Optional[str]:
    """
    Inspect an existing Excel file and decide whether it's a Building
    Survey or a Land Survey — or neither.

    Returns one of: "building", "land", or None (not recognizable).

    Recognition uses two independent signals so a renamed sheet or a
    one-cell tweak doesn't break detection:
      - active sheet title contains "Building Survey" / "Land Survey"
      - the distinguishing column in row 3 (K3) has the expected header
        text: "Total Building Area" for buildings, "Land Area" for land

    Either signal is sufficient; both being present is more confidence.
    """
    if not target_path.exists():
        return None
    try:
        wb = load_workbook(str(target_path), read_only=True, data_only=True)
        ws = wb.active
        title = (ws.title or "").lower()
        k3 = ws["K3"].value
        k3_str = "" if k3 is None else str(k3).lower()
    except Exception as e:
        log.warning("Could not inspect %s: %s", target_path, e)
        return None

    building_hits = 0
    land_hits = 0
    if "building" in title:
        building_hits += 1
    if "land" in title:
        land_hits += 1
    if "total building area" in k3_str:
        building_hits += 1
    if "land area" in k3_str:
        land_hits += 1

    if building_hits > land_hits and building_hits > 0:
        return "building"
    if land_hits > building_hits and land_hits > 0:
        return "land"
    return None


def _find_next_data_row(
    ws: Worksheet, columns: list[tuple[str, str]],
) -> tuple[int, int]:
    """
    Find the first empty data row in an append-target sheet and the
    next sequential `#` value to assign.

    Returns (next_row, next_record_number).

    "Occupied" means any of the schema data columns (C onwards in the
    template) holds a non-blank value. Column B (the # column) is NOT
    counted because the template pre-fills rows 7-14 with placeholder
    # values that have no associated data — counting them would make us
    skip those rows during append.

    The record-number continuation comes from the largest integer found
    in column B across rows we actually consider occupied. If that
    yields nothing, fall back to (next_row - FIRST_DATA_ROW + 1) so a
    fully-empty file numbers from 1.
    """
    last_occupied = SCAN_FIRST_ROW - 1
    max_b_value = 0
    # Cap the scan at ws.max_row + 2 so we always see one row past the
    # apparent end, but never run away on a sheet openpyxl thinks has
    # 1,048,576 rows because of dropdown ranges.
    scan_end = max(SCAN_FIRST_ROW, min(ws.max_row + 2, SCAN_FIRST_ROW + 5000))
    # Only consider schema data columns (C onwards). The template's
    # column B holds the auto-numbered #, which is pre-populated in
    # many empty rows of the template and would cause false positives.
    data_col_letters = [c for _, c in columns]
    for row in range(SCAN_FIRST_ROW, scan_end + 1):
        row_has_data = False
        for col in data_col_letters:
            v = ws[f"{col}{row}"].value
            if v not in (None, ""):
                row_has_data = True
                break
        if row_has_data:
            last_occupied = row
            b_val = ws[f"B{row}"].value
            if isinstance(b_val, int):
                max_b_value = max(max_b_value, b_val)
            elif isinstance(b_val, str) and b_val.strip().isdigit():
                max_b_value = max(max_b_value, int(b_val.strip()))

    next_row = last_occupied + 1
    next_number = max_b_value + 1 if max_b_value > 0 else (
        next_row - SCAN_FIRST_ROW + 1)
    return next_row, next_number


def _shift_validations_after_delete(ws: Worksheet, deleted_row: int) -> None:
    """
    openpyxl does not adjust data-validation ranges when rows are deleted,
    so after deleting `deleted_row` we rewrite every validation sqref:
    any row number strictly greater than the deleted row is decremented by
    one. This keeps the locked dropdowns aligned with the shifted cells.
    """
    import re

    def _fix_token(tok: str) -> str:
        # A token is like "S6:S1048576", "S4", or "S6". Shift each cell ref.
        def _fix_ref(ref: str) -> str:
            m = re.match(r"^(\$?[A-Z]+\$?)(\d+)$", ref)
            if not m:
                return ref
            col, rownum = m.group(1), int(m.group(2))
            if rownum > deleted_row:
                rownum -= 1
            return f"{col}{rownum}"
        return ":".join(_fix_ref(r) for r in tok.split(":"))

    for dv in ws.data_validations.dataValidation:
        parts = str(dv.sqref).split()
        dv.sqref = " ".join(_fix_token(p) for p in parts)


def _prepare_hint_row(ws: Worksheet, last_col_index: int) -> int:
    """
    Make the sheet ready for fresh data with no blank gap above it.

    The source template has:
      row 3: column headers
      row 4: a filled-in "Example" sample record
      row 5: the input-options hint row
      row 6+: the dropdown-validated data area

    We want the deliverable to start data immediately under the headers with
    NO example row and NO hint row left behind (clearing them only blanked
    the cells, which left a visible empty row above the data).

    Strategy:
      1. Extend each data-validation range that starts at row 6 down to
         row 5, so rows 5 (and after our deletes, the new data rows) keep
         their dropdowns.
      2. Physically DELETE the example row (row 4) and the hint row (row 5),
         shifting validation ranges up by one for each deletion (openpyxl
         doesn't do this automatically).

    After deleting two rows, the first data row is row 4 (was row 6). Returns
    that new first-data row so callers write to the right place.
    """
    # 1. Extend validation ranges from row 6 down to row 5.
    for dv in ws.data_validations.dataValidation:
        parts = str(dv.sqref).split()
        new_parts = []
        for p in parts:
            if ":" in p:
                left, right = p.split(":", 1)
                col_chars = "".join(c for c in left if c.isalpha())
                row_chars = "".join(c for c in left if c.isdigit())
                if row_chars == "6" and right.startswith(col_chars):
                    p = f"{col_chars}5:{right}"
            new_parts.append(p)
        dv.sqref = " ".join(new_parts)

    # 2. Delete the hint row (row 5) first, then the example row (row 4),
    #    shifting validations up after each. Deleting the higher row number
    #    first keeps the lower index valid.
    ws.delete_rows(HINT_ROW, 1)        # remove row 5 (hint)
    _shift_validations_after_delete(ws, HINT_ROW)
    ws.delete_rows(EXAMPLE_ROW, 1)     # remove row 4 (example)
    _shift_validations_after_delete(ws, EXAMPLE_ROW)

    # Data now starts where the example row used to be (row 4).
    return EXAMPLE_ROW


def _clear_leftover_numbers(ws: Worksheet, first_empty_row: int) -> None:
    """
    The template pre-fills column B (the "#") with 1, 2, 3, ... for many
    rows. After we write N records, the rows below still carry these stray
    numbers, which look like empty-but-numbered rows in the deliverable.
    Blank column B from the first empty row down to the last pre-numbered
    row so the output ends cleanly after the real data.
    """
    row = first_empty_row
    # Stop at the first row whose B is already blank (and not a leftover).
    # Cap the scan so we never run away.
    for row in range(first_empty_row, first_empty_row + 5000):
        b = ws.cell(row=row, column=2)
        v = b.value
        is_number = isinstance(v, int) or (
            isinstance(v, str) and v.strip().isdigit())
        if not is_number:
            break
        b.value = None


def _reset_sheet_view(ws: Worksheet) -> None:
    """
    Reset the worksheet's viewport so the data is actually visible when the
    file is opened.

    The KBC templates were saved with the view scrolled to column U and in
    Page Break Preview mode. A file written from such a template *contains*
    all the data, but Excel opens it scrolled past columns A-T (where the
    address/city/state data lives) and in a preview mode — so the sheet
    looks blank even though it is not.

    This forces:
      - Normal view (not pageBreakPreview / pageLayout)
      - the top-left cell back to A1
      - the active selection back to A1
    """
    sv = ws.sheet_view
    sv.view = "normal"             # was "pageBreakPreview"
    sv.topLeftCell = "A1"          # was "U1" — scroll back to the start
    # Reset the active cell/selection to A1 so nothing is mid-sheet.
    try:
        from openpyxl.worksheet.views import Selection
        sv.selection = [Selection(activeCell="A1", sqref="A1")]
    except Exception:
        pass


def _coerce_for_excel(field_name: str, value):
    """
    Final pre-write cleanup. Pydantic gives us mostly-clean values; this
    function handles the last-mile conversions Excel cares about.
    """
    if value is None:
        return None
    if field_name == "state":
        return normalize_state(value if isinstance(value, str) else str(value))
    if field_name == "date_available":
        d = normalize_date(value)
        return d if d else None
    if field_name == "office_space_sf":
        # Special case: this field accepts a number OR the literal "BTS".
        if isinstance(value, str) and value.strip().upper() == "BTS":
            return "BTS"
        n = clean_number(value)
        return n if n is not None else value
    # Generic numeric fields: try to coerce strings to floats.
    if field_name in {
        "total_building_area_sf", "available_space_sf", "land_area_acres",
        "year_built", "clear_height_min_ft", "clear_height_max_ft",
        "dock_doors", "grade_level_doors", "column_spacing_width_ft",
        "column_spacing_depth_ft", "auto_parking_spaces", "trailer_parking_spaces",
        "annual_asking_rate_psf", "sale_price", "estimated_annual_opex_psf",
        "latitude", "longitude",
    }:
        n = clean_number(value)
        return n if n is not None else None
    return value


def _write_record(
    ws: Worksheet,
    row: int,
    record_data: dict,
    confidence: dict[str, float],
    columns: list[tuple[str, str]],
    low_conf_threshold: float,
    flagged: list[str],
    row_number: int,
) -> None:
    """Write a single record at `row`. Leaves None fields blank."""
    # Auto-numbered # column
    ws[f"B{row}"] = row_number

    for field_name, col in columns:
        value = record_data.get(field_name)
        cleaned = _coerce_for_excel(field_name, value)
        if cleaned is None or cleaned == "":
            continue
        cell = ws[f"{col}{row}"]
        cell.value = cleaned
        # Highlight if the LLM flagged this field as low-confidence.
        if confidence.get(field_name, 1.0) < low_conf_threshold:
            cell.fill = LOW_CONFIDENCE_FILL
            flagged.append(f"{col}{row}")


# ---------------------------------------------------------------------------
# Public writers
# ---------------------------------------------------------------------------

def write_building_records(
    records: list[BuildingRecord],
    cfg: Config,
    output_dir: Optional[Path] = None,
    target_path: Optional[Path] = None,
    output_name: Optional[str] = None,
) -> WriteSummary:
    """
    Write building records to Excel.

    Two modes, mutually exclusive:
      - target_path is None  -> create a fresh file in output_dir (or
                                cfg.default_output_dir). The filename
                                comes from output_name if given,
                                otherwise the legacy timestamped form.
                                This is the legacy behavior.
      - target_path is set   -> open that existing file and APPEND the
                                records after the last occupied row.
                                The file's survey type must match
                                (caller should have validated this).
                                output_name is ignored.

    In append mode we skip the hint-row prep and viewport reset — both
    were already done the first time the file was written. We just
    find the next free row, continue the # numbering from there, and
    write rows on top of an unmodified file.
    """
    appending = target_path is not None
    if appending:
        out_path = target_path
        if not out_path.exists():
            raise FileNotFoundError(
                f"Append target does not exist: {out_path}")
    else:
        out_dir = output_dir or Path(cfg.default_output_dir)
        out_path = _prepare_output("building", out_dir, output_name=output_name)

    wb = load_workbook(str(out_path))
    ws = wb.active

    if appending:
        start_row, start_number = _find_next_data_row(ws, BUILDING_COLUMNS)
    else:
        # Remove the example + hint rows entirely (no blank gap above data),
        # preserving dropdowns; data starts at the returned row.
        start_row = _prepare_hint_row(ws, HINT_ROW_LAST_COL_BUILDING)
        # Reset the viewport so the data is visible on open (template opens
        # scrolled to column U in Page Break Preview).
        _reset_sheet_view(ws)
        start_number = 1

    flagged: list[str] = []
    for i, rec in enumerate(records):
        row = start_row + i
        _write_record(
            ws=ws,
            row=row,
            record_data=rec.model_dump(),
            confidence=rec.confidence_notes,
            columns=BUILDING_COLUMNS,
            low_conf_threshold=cfg.low_confidence_threshold,
            flagged=flagged,
            row_number=start_number + i,
        )

    # Clear any stray pre-numbered "#" values in the rows below the data.
    _clear_leftover_numbers(ws, start_row + len(records))

    wb.save(str(out_path))
    return WriteSummary(output_path=out_path, rows_written=len(records),
                        flagged_cells=flagged, appended=appending,
                        start_row=start_row)


def write_land_records(
    records: list[LandRecord],
    cfg: Config,
    output_dir: Optional[Path] = None,
    target_path: Optional[Path] = None,
    output_name: Optional[str] = None,
) -> WriteSummary:
    """
    Write land records to Excel. Same fresh-vs-append semantics as
    write_building_records — see its docstring.
    """
    appending = target_path is not None
    if appending:
        out_path = target_path
        if not out_path.exists():
            raise FileNotFoundError(
                f"Append target does not exist: {out_path}")
    else:
        out_dir = output_dir or Path(cfg.default_output_dir)
        out_path = _prepare_output("land", out_dir, output_name=output_name)

    wb = load_workbook(str(out_path))
    ws = wb.active

    if appending:
        start_row, start_number = _find_next_data_row(ws, LAND_COLUMNS)
    else:
        # Remove the example + hint rows entirely (no blank gap above data),
        # preserving dropdowns; data starts at the returned row.
        start_row = _prepare_hint_row(ws, HINT_ROW_LAST_COL_LAND)
        # Reset the viewport so the data is visible on open.
        _reset_sheet_view(ws)
        start_number = 1

    flagged: list[str] = []
    for i, rec in enumerate(records):
        row = start_row + i
        _write_record(
            ws=ws,
            row=row,
            record_data=rec.model_dump(),
            confidence=rec.confidence_notes,
            columns=LAND_COLUMNS,
            low_conf_threshold=cfg.low_confidence_threshold,
            flagged=flagged,
            row_number=start_number + i,
        )
        # Add the per-acre formula =Q/K only when we have both numbers.
        # If either is missing the cell stays blank (no #DIV/0!).
        if rec.sale_price is not None and rec.land_area_acres:
            ws[f"P{row}"] = f"=Q{row}/K{row}"

    # Clear any stray pre-numbered "#" values in the rows below the data.
    _clear_leftover_numbers(ws, start_row + len(records))

    wb.save(str(out_path))
    return WriteSummary(output_path=out_path, rows_written=len(records),
                        flagged_cells=flagged, appended=appending,
                        start_row=start_row)
